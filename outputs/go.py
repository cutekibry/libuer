#!/usr/bin/python3

import openpyxl
import csv


# getdata
def getdata(file):
    f = open(file, 'r', encoding='utf-8')
    reader = csv.reader(f)
    data = []
    for row in list(reader)[1:]:
        data.append([int(row[1]), row[2], row[9], row[10]])
    return data


iotdata = getdata('iot.csv')
cmdata = getdata('cutekibry.csv') + getdata('moonoshawott.csv')
cmdata = sorted(cmdata, key=lambda x: x[3], reverse=True)


# create
book = openpyxl.Workbook()

book.remove(book.active)

iot = book.create_sheet()
iot.title = 'iot'
cm = book.create_sheet()
cm.title = 'cutekibry & moonoshawott'


# write
for i, x in enumerate(['题目编号', '题目名称', '提交者', '提交时间'], start=1):
    cm.cell(1, i, x)
for i, x in enumerate(cmdata, start=2):
    for j, y in enumerate(x, start=1):
        cm.cell(i, j, y)

for j, x in enumerate(['比较', '题目编号', '题目名称', '提交者', '提交时间'], start=1):
    iot.cell(1, j, x)
for i, x in enumerate(iotdata, start=2):
    iot.cell(
        i, 1, "=IF(COUNTIF('cutekibry & moonoshawott'!$A$2:$A$1000, B{}), \"\", \"X\")".format(i))
    for j, y in enumerate(x, start=2):
        iot.cell(i, j, y)

n = iot.max_row
iot.cell(n + 1, 1, "Total:")
iot.cell(n + 1, 2, n - 1)
iot.cell(n + 2, 1, "Todo:")
iot.cell(n + 2, 2, "=COUNTIF(A2: A%d, \"X\")" % n)



# save
book.save('iot-cmp.xlsx')
